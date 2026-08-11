import requests
import pandas as pd
import shutil
from openpyxl import load_workbook

URL = "https://hydro-back.imgw.pl/map/stations/meteorologic?onlyMainStations=false"
EXCEL_FILE = "all_stations.xlsx"


def fetch_imgw_stations():
    response = requests.get(
        URL,
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=10
    )
    response.raise_for_status()

    data = response.json()
    stations = data.get("stations", [])

    return {
        str(station.get("id")): {
            "name": station.get("n"),
            "lon": station.get("lo"),
            "lat": station.get("la"),
        }
        for station in stations
        if station.get("id")
    }


def build_header_map(ws):
    headers = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    for index, cell in enumerate(header_row, start=1):
        if cell.value is None:
            continue
        headers[str(cell.value).strip().lower()] = index
    return headers


def find_column(headers, candidates, default):
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
    return default


def normalize_status(value):
    if value is None:
        return ""
    return str(value).strip().upper()


def main():
    print("📄 Wczytywanie Excel...")
    df = pd.read_excel(EXCEL_FILE, dtype={"Station_id": str})
    existing_ids = set(df["Station_id"].dropna())

    print("🌐 Pobieranie stacji z IMGW...")
    imgw_stations = fetch_imgw_stations()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = build_header_map(ws)

    id_col = find_column(headers, ["station_id", "stationid", "id"], 1)
    name_col = find_column(headers, ["name", "station_name", "stationname", "nazwa"], 2)
    coords_col = find_column(headers, ["coords", "coordinates", "geo", "location", "wspolrzedne"], 3)
    extra_col = find_column(headers, ["notes", "uwagi", "comment", "remarks"], 4)
    status_col = find_column(headers, ["status", "station_status", "stan", "state"], 5)

    missing_ids = set(imgw_stations.keys()) - existing_ids
    closed_rows = []

    for row in ws.iter_rows(min_row=2):
        sid_cell = row[id_col - 1].value
        if sid_cell is None:
            continue

        sid = str(sid_cell).strip()
        if sid not in imgw_stations:
            continue

        status_value = row[status_col - 1].value if status_col <= len(row) else None
        if normalize_status(status_value) == "CLOSED":
            closed_rows.append((sid, row))

    if not missing_ids and not closed_rows:
        print("Brak nowych stacji ani zamkniętych stacji do aktywacji.")
        return

    backup_file = EXCEL_FILE.replace(".xlsx", "_backup.xlsx")
    shutil.copy(EXCEL_FILE, backup_file)
    print(f"📦 Backup zapisany jako: {backup_file}")

    activated_count = 0
    for sid, row in closed_rows:
        s = imgw_stations[sid]
        coords_str = f"[{s['lon']}, {s['lat']}]"

        row[id_col - 1].value = sid
        row[name_col - 1].value = s["name"]
        row[coords_col - 1].value = coords_str
        row[extra_col - 1].value = None
        row[status_col - 1].value = "ACTIVE"

        activated_count += 1
        print(f"{sid} | {s['name']} | {coords_str} -> ACTIVE")

    added_count = 0
    for sid in sorted(missing_ids, key=lambda x: int(x) if x.isdigit() else x):
        s = imgw_stations[sid]
        coords_str = f"[{s['lon']}, {s['lat']}]"

        print(f"{sid} | {s['name']} | {coords_str}")
        ws.append([
            sid,
            s["name"],
            coords_str,
            None,
            "ACTIVE"
        ])
        added_count += 1

    wb.save(EXCEL_FILE)

    print(f"\n✅ Zaktualizowano {activated_count} zamkniętych stacji.")
    print(f"✅ Dodano {added_count} nowych stacji do {EXCEL_FILE}")
    print("📐 Szerokość kolumn została zachowana.")


if __name__ == "__main__":
    main()